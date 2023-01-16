from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT as NavigationToolbar
from PyQt5.QtWidgets import QFileDialog
from math import *
import os
import numpy as np
import pandas as pd
import time
import shutil
import openpyxl
from variables import *
from datetime import datetime
# from scipy.spatial.distance import euclidean

def GET_XY_FROM_MOUSE(Table):
    # Get cells value has been selected
    selected = Table.selectedIndexes()
    # Intial empty lists
    X, Y, selected_list_index = [],[], []

    # Check selected items are empty or not
    if len(selected) != 0: selected_list_index = [([ix.column(), ix.row()]) for ix in selected] # if not
    else: 
        return [], [], [], 0 #Return X, Y, status if yes
        
    X_index, Y_index = [], []

    X_index = [(selected_list_index[i][0]) for i in range(0, len(selected_list_index))]
    Y_index = [(selected_list_index[i][1]) for i in range(0, len(selected_list_index))]
    
    X_index = list(set(X_index)) # Column index
    Y_index = list(set(Y_index)) # Row index
    row_start_index, row_end_index = Y_index[0], Y_index[len(Y_index)-1]

    if (len(X_index) <= 1) or (len(Y_index) < 1): 
        return X, Y, [], 0 #Return X, Y, (row start, row end index) and status
       
    X_coordinates, Y_coordinates = [], []
    X_coordinates = [([Y_index[i], X_index[0]]) for i in range(0,len(Y_index))]
    Y_coordinates = [([Y_index[i], X_index[1]]) for i in range(0,len(Y_index))]
    
    try:
        X = [(float(Table.item(x[0],x[1]).text().replace(",",""))) for x in X_coordinates]
        Y = [(float(Table.item(y[0],y[1]).text().replace(",",""))) for y in Y_coordinates]
    except (ValueError, AttributeError):
        return [], [], [-1, -1], 0 # X, Y, (row start, row end index) and status

    return X, Y, [row_start_index, row_end_index], 1 # X, Y, (row start, end index), status
# End of GET_XY_FROM_MOUSE function #

def SCATTER_CHART(X_list, Y_list, chart, NC_list, NC_split_char, Pad_name=[], Pad_number=[]):
    # Add chart in tab1
    chart.axes.axhline(y=0, color='#FFFFFF', linestyle='-.', linewidth=0.5)
    chart.axes.axvline(x=0, color='#FFFFFF', linestyle='-.', linewidth=0.5)
    
    NC_pos = []
    Probe_pos = []
    First_pad_name = ""

    if len(Pad_name) > 0 and len(Pad_number) > 0: 
        if Pad_name[0] != "" and Pad_number[0] != 0: 
            First_pad_name = "Pad#" + Pad_number[0] + "; " + Pad_name[0]
    if len(Pad_name) > 0 and len(Pad_number) == 0: 
        if Pad_name[0] != "":
            First_pad_name = Pad_name[0]
    if len(Pad_name) == 0 and len(Pad_number) > 0: 
        if Pad_number[0] != 0: 
            First_pad_name = "Pad#" + Pad_number[0]

    if len(X_list) > 0:
        if len(NC_list) > 0 and NC_split_char != "":
            NC_pos = [i for i in range(len(NC_list)) if NC_list[i] == NC_split_char]
            Probe_pos = [i for i in range(len(NC_list)) if NC_list[i] != NC_split_char]

            X_list_NC, Y_list_NC, Annotate_NC = [], [], []
            
            X_list_NC = [X_list[i] for i in NC_pos]
            Y_list_NC = [Y_list[i] for i in NC_pos]
        
            X_list_probe, Y_list_probe =[], []
            
            [X_list_probe.append(X_list[i]) for i in Probe_pos]
            [Y_list_probe.append(Y_list[i]) for i in Probe_pos]

            chart.axes.scatter(X_list_NC, Y_list_NC, color = "white", marker="s", s=10, label="NC")
            chart.axes.scatter(X_list_probe, Y_list_probe, color = "orange", marker="o", s=10, label="Probe")

            if len(Pad_name) > 0 and len(Pad_number) > 0: 
                if Pad_name[Probe_pos[0]] != "" and Pad_number[Probe_pos[0]] != "":
                    First_pad_name = "Pad#" + Pad_number[Probe_pos[0]] + "; " + Pad_name[Probe_pos[0]]
            if len(Pad_name) > 0 and len(Pad_number) == 0: 
                if Pad_name[Probe_pos[0]] != "":
                    First_pad_name = Pad_name[Probe_pos[0]]
            if len(Pad_name) == 0 and len(Pad_number) > 0: 
                if Pad_number[Probe_pos[0]] != "":
                    First_pad_name = "Pad#" + Pad_number[Probe_pos[0]]
            if First_pad_name != "":
                chart.axes.annotate(First_pad_name, 
                                    xy=(X_list_probe[0],Y_list_probe[0]), 
                                    color="orange")
        else:
            chart.axes.scatter(X_list, Y_list, color = "orange", marker="o", s=3, label="Probe")
            if First_pad_name != "":
                chart.axes.annotate(First_pad_name, 
                                    xy = (X_list[0], Y_list[0]), 
                                    color = "white")

        chart.axes.axis('equal')
        chart.axes.set_xlim([min(X_list)-5, max(X_list)+5]) 
        chart.axes.set_ylim([min(Y_list)-5, max(Y_list)+5])
        chart.axes.legend(facecolor="grey", labelcolor="w", frameon=True)

    chart.axes.set_xlabel('X-axis')
    chart.axes.set_ylabel('Y-axis')

    chart.axes.xaxis.label.set_color('#FFFFFF')
    chart.axes.tick_params(axis='x', colors='#FFFFFF')
    chart.axes.yaxis.label.set_color('#FFFFFF')
    chart.axes.tick_params(axis='y', colors='#FFFFFF')

    chart.axes.spines['bottom'].set_color('#FFFFFF')
    chart.axes.spines['top'].set_color(None)   
    chart.axes.spines['left'].set_color('#FFFFFF')
    chart.axes.spines['right'].set_color(None)  

    chart.axes.patch.set_facecolor('None')
    chart.axes.patch.set_alpha(0.0)

    chart.setStyleSheet("background-color:transparent;")
    
    # Add toolbar chart
    toolbar = NavigationToolbar(chart)
    toolbar.setStyleSheet("background-color: rgb(255,255,255);"
                        "border-radius: 8px;"
                        "color: black;"
                        )
    
    return chart, toolbar
# End of SCATTER_CHART function #

def CAL_DIE_SIZE(X, Y):
    # return Size in X and Size in Y
    return (max(X) - min(X)), (max(Y)-min(Y))
# End of CAL_DIE_SIZE function #

#  Start of ClosestPair class
class ClosestPair:
    def __init__(self, X, Y):
        super().__init__()
        self.min_distance = 0
        self.P = [(X[i], Y[i]) for i in range(len(X))]
        
        n = len(self.P)
        Px = sorted(self.P, key=lambda x: x[0])
        Py = sorted(self.P, key=lambda x: x[1])
        
        self.min_distance = self.Calculate(Px, Py, n)
        
    def Calculate(self, Px, Py, n):
        try:
            # Simple case
            if n == 2:
                return dist(self.P[0], self.P[1])
            if n == 3:
                return min(dist(self.P[0], self.P[1]), dist(self.P[0], self.P[2]), dist(self.P[1], self.P[2]))
        except:
            pass
        
        # Divide 
        mid = n // 2
        d_left = self.Calculate(Px, Py[:mid], mid)
        d_right = self.Calculate(Py, Py[mid:], n - mid)
        
        min_distance = min(d_left, d_right)
        
        S = [] # points list in strip
        for p in Px:
            if abs(p[0] - Px[mid][0]) < min_distance:
                S.append(p)
                
        min_distance_in_strip = float('inf')
        for i in range(min(6, len(S) - 1), len(S)):
            for j in range(max(0, i - 6), i):
                current_dis = dist(S[i], S[j])
                if current_dis < min_distance:
                    min_distance_in_strip = current_dis
        
        return min(min_distance, min_distance_in_strip)

    def min_distance_in_points(self):
        return self.min_distance
# End of ClosestPair class

def GET_FILE(): # No input
    filter_ = "Excel File (*.xlsx *xlsm)"
    file_path = QFileDialog.getOpenFileName(caption='selectFile',
                                            directory="",
                                            filter=filter_,
                                            initialFilter="")
    file_name = file_path[0].split('/')[-1]
    # return:
    # file_path, file_name and file_extension
    return file_path[0], file_name, os.path.splitext(file_name)[1]
# End of GET_FILE function

def GET_SPEC_FILE(): # No input
    filter_ = "Txt File (*.txt)"
    file_path = QFileDialog.getOpenFileName(caption='selectFile',
                                            directory="",
                                            filter=filter_,
                                            initialFilter="")
    file_name = file_path[0].split('/')[-1]

    return file_path[0], file_name, os.path.splitext(file_name)[1]
# End of GET_FILE function

def DIE_PATTERN(X, Y, config_table, Stepping_distance):
    new_X = []
    new_Y = []
    config_array = np.array([])
    return new_X, new_Y

def UPDATE_PARAMETER_FOR_EXISTED_PROJECT(card_part_number):
    spec_file_name = card_part_number + "_spec_file.txt"
    return spec_file_name

def GET_FILE_PATH(path_file, file_need_find):
    path_file_open = open(path_file, 'r')
    content = path_file_open.readlines()
    path_file_open.close()

    file_name, file_extension = "", ""

    for i in content:
        if file_need_find in i:
            file_name = i[len(file_need_find+"="):i.index(".")]
            if "." in i:
                file_extension = i[i.index("."):]
                file_extension = file_extension[1:]
            break
        
    file_extension = file_extension.replace("\n", "")
    file_name = file_name.replace("\n", "")
    # completed to get file name, file extension from path.txt file
    return file_name, file_extension

def GET_FOLDER_PATH(path_file, folder_name_to_find):
    # Open paths.txt file
    path_file_open = open(path_file, 'r')
    # Read paths.txt file and store file content to content variable
    content = path_file_open.readlines()
    # Close file
    path_file_open.close()
    # Inital variable
    folder_path = ""
    # Find file name string in content file
    for i in content:
        if folder_name_to_find in i:
            folder_path = i[len(folder_name_to_find+"="):]
            break
    folder_path = folder_path.replace('\n', "")
    # completed to get folder path
    return folder_path

def GET_REVISION_STRING(path_file, string_to_find):
    # Open paths.txt file
    path_file_open = open(path_file, 'r')
    # Read paths.txt file and store file content to content variable
    content = path_file_open.readlines()
    # Close file
    path_file_open.close()
    # Inital variable
    revision_string = ""
    # Find file name string in content file
    for i in content:
        if string_to_find in i:
            revision_string = i[len(string_to_find)+len("="):i.index("\n")]
            break
                
    # completed to get folder path
    return revision_string

# Start of REVISION_CHANGE_TYPE function:
# Input:
# - old_revision: old revision
# - type: Edit, Release, Override or New
def REVISION_CHANGE_TYPE(old_revision, type = "Edit"): 
    # Check old_revision string if it is empty 
    if old_revision == "":
        return "" 
    
    # Keep revision if type is Override
    if type == "Override":
        return old_revision

    # If old_revision string is not empty:
    # Initial variables
    numeric_revision = "" # from 00 to 99, example: 01, 02, 03. Note: 00 is used to initial
    alpha_revision = "" # from A to Z, AA to ZZ, example: A, B, AA, AB
    mix_revision = "" # include A and 00 or AA and 00, example: A01, AB01
    new_revision = ""

    # Check numberic or alpha or mix type:
    if old_revision.isdigit() == True: # revision value is like "01", "02", or "03"
        numeric_revision = old_revision
        alpha_revision = ""
        mix_revision = ""

    elif old_revision.isalpha() == True: # revision value is like "A" or "AA"
        alpha_revision = old_revision
        numeric_revision = ""
        mix_revision = ""

    else: # Revision value is like "A01" or "AA01"
        mix_revision = old_revision
        alpha_revision = ""
        numeric_revision = ""

    s = [] # Convert revision string into list, like s = ['0','1'] or ['A']
    if numeric_revision != "":
        s = [i for i in numeric_revision]
    elif alpha_revision != "":
        s = [i for i in alpha_revision]
    elif mix_revision != "":
        s = [i for i in mix_revision]
    
    if type == "Edit":
        if s: # if list is not empty
            s.reverse()

            # Case 1: digit, change from 01 to 02, keyword: digit2digit
            if numeric_revision != "" and alpha_revision == "" and mix_revision == "":
                for i, element in enumerate(s):
                    if ord(element) >= 48 and ord(element) < 57:
                        s[i] = chr(ord(element) + 1)
                        break
                    else:
                        s[i] = "0"
            # end if numeric_revision != "" and alpha_revision == "" and mix_revision == "" #

            # Case 2: alpha, change from A to A01, keyword: alpha2alphadigit
            if numeric_revision == "" and alpha_revision != "" and mix_revision == "":
                s.reverse()
                s.append("01")
                s.reverse()
            # end if numeric_revision == "" and alpha_revision != "" and mix_revision == ""
           
            # Case 3: alpha + digit,  change from A01 to A02, or AA01 to AA02, keyword: alphadigit2alphadigit
            if numeric_revision == "" and alpha_revision == "" and mix_revision != "":
                if len(mix_revision) < 3:
                    return ""
                # end if len(mix_revision) < 3

                if len(mix_revision) == 3:
                    alpha = mix_revision[0]
                    number = [i for i in mix_revision[1:]]
                    number.reverse()
                    for i, element in enumerate(number):
                        if ord(element) >= 48 and ord(element) < 57: # in ACSII, CHAR "A" = 48 in DEC, "Z" = 57
                            number[i] = chr(ord(element) + 1)
                            break
                        else:
                            number[i] = "0"
                    number_ = ""
                    for i in number:
                        number_ = number_ + i
                    number_ = number_ + alpha
                    
                    s = []
                    s = [i for i in number_]
                # end if len(mix_revision) == 3

                if len(mix_revision) == 4:
                    alpha = [i for i in mix_revision[:2]]
                    number = [i for i in mix_revision[2:]]
                    number.reverse()
                    for i, element in enumerate(number):
                        if ord(element) >= 48 and ord(element) < 57:
                            number[i] = chr(ord(element) + 1)
                            break
                        else:
                            number[i] = "0"
                    
                    number_ = ""
                    alpha.reverse()
                    number.reverse()
                    for i in number:
                        number_ = number_ + i
                    for i in alpha:
                        number_ = i + number_
                    
                    s = []
                    s = [i for i in number_]
                    s.reverse()
                # end if len(mix_revision) == 4

            s.reverse()
            new_revision = ""
            for i in s:
                new_revision = new_revision + i
        else:
            return ""
        
        # end if s:
    # end if type == "Edit"

    if type == "Release":
        if s:
            s.reverse()
            # Case 1: digit, example: 01 to A, keyword: digit2alpha
            if numeric_revision != "" and alpha_revision == "" and mix_revision == "":
                s = []
                s.append("A")

            # Case 2: mix, example: A01 to B, keyword: alphadigit2alpha
            if numeric_revision == "" and alpha_revision == "" and mix_revision != "":
                alpha   = []
                number  = []

                if len(mix_revision) == 3:
                    alpha.append(mix_revision[0])
                    number  = [i for i in mix_revision[1:]]

                if len(mix_revision) == 4:
                    alpha   = [i for i in mix_revision[:2]]
                    number  = [i for i in mix_revision[2:]]
                
                # Example: A01 to A
                if len(alpha) == 1:
                    if ord(alpha[0]) >= 65 and ord(alpha[0]) < 90:
                        alpha[0] = chr(ord(alpha[0]) + 1)
                    else:
                        alpha[0] = "AA"
                if len(alpha) == 2:
                    alpha.reverse()
                    for i, e in enumerate(alpha):
                        if ord(e) >= 65 and ord(e) < 90:
                            alpha[i] = chr(ord(alpha[0]) + 1)
                            break
                        else:
                            alpha[0] = "A"
                    alpha.reverse()
                s = []
                for i in alpha:
                    s.append(i)
                s.reverse()
            # Case 3: alpha, example: A to B, keyword: 
            if numeric_revision == "" and alpha_revision != "" and mix_revision == "":
                # Example: A to B, Z to AA
                if alpha_revision == "Z":
                    return "AA"
                if alpha_revision == "ZZ":
                    return "AAA"
                
                for i, element in enumerate(s):
                    if ord(element) >= 65 and ord(element) < 90:
                        s[i] = chr(ord(element) + 1)
                        break
                    else:
                        s[i] = "A"
            s.reverse()
            new_revision = ""
            for i in s:
                new_revision = new_revision + i
        else:
            return ""
        # end if s
    # if type == "Release"
    return new_revision # like: 02, A, B, A02

def EXPORT_XY_FORMAT_FOR_IUA_PLUS_FILE(X, Y, \
                                        pad_number_list, \
                                        pad_name_list, \
                                        xy_input_unit,  \
                                        status=[], \
                                        card_part_number="", \
                                        sheet_name = 'Sheet1'):
    # Input Parameters:
    #   1. table_TableWidget is a TableWidget
    #   3.card_part_number as PCX-000000, MSP-000000
    #   4.status: Release, Edit, Override, New
    #   5.sheet_name is "Sheet1" as default
    # Note Unit is mm as default, convert to mm if unit is not 'mm'
    
    if not X or not Y: # not export anything of X and Y list are empty
        # print("empty")
        return "Empty"
        
    # Get file name, file extension from path.txt file
    # example: file_name = "YYY-XXXXXX_XY input format for IUA_plus_Rev00" and file_extension = "xlsx"
    file_name, file_extension = GET_FILE_PATH('paths.txt', 'XY_FORMAT_FOR_IUA_PLUS_FILENAME')
    
    path = GET_FOLDER_PATH('paths.txt', 'design_folder_path') + "/" + str(card_part_number) + '-xx'
    
    file_exist = False
    XY_exist_name = ""
    
    if card_part_number != "":
 
        if os.path.exists(path):
            dir_list = os.listdir(path)

            for i in dir_list:
                
                string = card_part_number + "-xx_XY Format for IUA Plus" # like: "PCX-000000-xx_XY Format for IUA Plus"

                if (i.find(string) != -1) and (i[i.find(".")+1:] == file_extension):
                    file_exist = True
                    XY_exist_name = i
                else:
                    continue
            # end for
        else:
            return "Card Folder doesn't exist!" # end if
    else:
        return "Card part number is empty!"  # alway need card_part_number to add to file, return empty char if this variable is ""
    
    if file_exist == False:
        # Get folder template path
        folder_template_location        = GET_FOLDER_PATH('paths.txt', 'folder_template_path')
        revision_string_format          = GET_REVISION_STRING('paths.txt', "REVISION_FORMAT") # like: "Rev"
        revision_pos_in_file_name       = file_name.index(revision_string_format) # position of "Rev" in file_name
        revision_string                 = file_name[revision_pos_in_file_name:] # like: "Rev00"
        new_file_path, new_file_name    = "", ""
        
        if card_part_number != "":
            new_file_name = file_name.replace("YYY-XXXXXX", card_part_number)
            new_file_name = new_file_name.replace(new_file_name[revision_pos_in_file_name:], "Rev01")
            new_file_path = path + "/" + new_file_name + "." + file_extension
        
        scr = folder_template_location + "/" + file_name + "." + file_extension
        try:
            shutil.copyfile(scr, new_file_path)
        except:
            # Create new file
            if not os.path.exists(new_file_path): 
                return "A folder does not exist"
            else:                       
                return "Error!"     
        # Load workbook
        workbook = openpyxl.load_workbook(new_file_path)
        # Active sheet "Sheet1"
        sheet = workbook[sheet_name]
        for row in range(len(X)):
            sheet.cell(row+2, 1).value = X[row] 
            sheet.cell(row+2, 2).value = Y[row]
            if len(pad_number_list) == len(X):
                sheet.cell(row+2, 3).value = pad_number_list[row]
            if len(pad_name_list) == len(X):
                sheet.cell(row+2, 4).value = pad_name_list[row]  
        workbook.save(new_file_path)
        return "Done!"
        # end if file_exist == False
        
    elif file_exist == True: # if the file doesn't exist, check status = ['Edit', 'Release', Override']

        file_name = XY_exist_name
        # Get folder template path
        folder_template_location    = GET_FOLDER_PATH('paths.txt', 'folder_template_path')
        # Get revision in file name:
        revision_string_format      = GET_REVISION_STRING('paths.txt', "REVISION_FORMAT") # like: "Rev"
        revision_pos_in_file_name   = file_name.index(revision_string_format) # a position
        revision_string             = file_name[revision_pos_in_file_name:] # like: "00" or "01"
      
        # Get old revision
        old_revision = revision_string[len(revision_string_format):revision_string.find(".")]

        if card_part_number != "": # if card_part_number is available.
            # status elements as following:
            #  status[0] ~ REVISION_STATUS[0] = "Edit"
            #  status[1] ~ REVISION_STATUS[1] = "Release"
            #  status[2] ~ REVISION_STATUS[2] = "Override"
            if status["Edit"] == True and status["Release"] == False and status["Override"] == False: # Edit
                # Get new revision base on status value
                new_revision = REVISION_CHANGE_TYPE(old_revision, REVISION_STATUS[0])
            elif status["Edit"] == False and status["Release"] == True and status["Override"] == False: # Release
                # Get new revision base on status value
                new_revision = REVISION_CHANGE_TYPE(old_revision, REVISION_STATUS[1])
            elif status["Edit"] == False and status["Release"] == False and status["Override"] == True: # Override
                # Get new revision base on status value
                new_revision = REVISION_CHANGE_TYPE(old_revision, REVISION_STATUS[2])
            
            # Get new file name and new file location
            new_file_name = file_name.replace(old_revision, new_revision)
            new_file_path = path + "/" + new_file_name
            
            # Get exist file name/file path
            scr = path + "/" + XY_exist_name
            
            # Copy file from scr to new_file_path and rename it
            shutil.copyfile(scr, new_file_path)
            
            # load workbook
            load_workbook = openpyxl.load_workbook(new_file_path)
            # Active "Sheet1" in workbook
            sheet = load_workbook[sheet_name]
            
            # Add data from X, Y list into actived sheet
            for row in range(len(X)):
                sheet.cell(row+2, 1).value = X[row] 
                sheet.cell(row+2, 2).value = Y[row]
                if len(pad_number_list) == len(X):
                    sheet.cell(row+2, 3).value = pad_number_list[row]
                if len(pad_name_list) == len(Y):
                    sheet.cell(row+2, 4).value = pad_name_list[row]  
            # Save workbook
            load_workbook.save(new_file_path)
        else: # if card_part_number is empty
            pass
        # end if
    else:
        return ""
    # Completed
    return "EXPORT_XY_FORMAT_FOR_IUA_PLUS_FILE"

def EXPORT_PCB_PAD_LOCATION_FILE():
    return "EXPORT_PCB_PAD_LOCATION_FILE"

def EXPORT_ARRAY_FULL_SITE_FOR_REFERENCE_FILE(X, Y, \
                                                pad_number_list, \
                                                pad_name_list, \
                                                dut_name_list, \
                                                xy_input_unit,  \
                                                status=[], \
                                                stepping_distance = [], \
                                                dut_name_format="", \
                                                card_part_number="", \
                                                sheet_name = 'XY list'):
    # Input parameter:
    #   1. X: X list after pattern
    #   2. Y: Y list after pattern
    #   3. pad_number_list: pad number list after pattern
    #   4. xy_input_unit: mm as default, convert to mm if not
    #   5. status: "Release", "Edit" or "Override"
    #   6. stepping_distance=[X_step, Y_step] to check with X, Y one DUT site.
    #   7. card_part_number
    #   8. sheet_name = "XY list" as default for this file
    # Return value:
    #   Return file name if this function are completed
    
    # Check X and Y list are empty or not
    if not X or not Y: # not export anything of X and Y list are empty
        # print("empty")
        return "Empty"
    
    # Declare some variables
    path = ""
    file_exist = False
    XY_exist_name = ""
    
    # Get file_name and file_extension from template
    file_name, file_extension = GET_FILE_PATH('paths.txt', 'ARRAY_FULL_SITE_TEMPLATE_FILENAME')
    
    # Check file does exited or not yet base on card_part_number
    if card_part_number != "":
        path = GET_FOLDER_PATH('paths.txt', 'design_folder_path') + "/" + str(card_part_number) + '-xx'
        if os.path.exists(path):
            dir_list = os.listdir(path)
            for i in dir_list:
                string = card_part_number + "-xx_Array full sites for reference" 
                # string is like: "PCX-000000-xx_Array full sites for reference"
                if (i.find(string) != -1) and (i[i.find(".")+1:] == file_extension):  #if string is included in files list and match file extension
                    file_exist = True
                    XY_exist_name = i # get exist file name
                else: # if string is not included in files list and not match file extension
                    continue
            # end for
        else:
            return "Card Folder doesn't exist!" # end if
    else:
        return "Card part number is empty!"  # alway need card_part_number to add to file, return empty char if this variable is ""
    
    if file_exist == False: # if file doesn't exist, creating new file Rev01 from template Rev00
        folder_template_location    = GET_FOLDER_PATH("paths.txt", 'folder_template_path')
        revision_string_format      = GET_REVISION_STRING('paths.txt', 'REVISION_FORMAT')
        revision_pos_in_file_name   = file_name.index(revision_string_format)
        revision_string             = file_name[revision_pos_in_file_name:] # example: Rev00 from template file
        new_file_path, new_file_name = "", ""
        
        if card_part_number != "":
            new_file_name = file_name.replace("YYY-XXXXXX", card_part_number)
            new_file_name = new_file_name.replace(new_file_name[revision_pos_in_file_name:], "Rev01")
            new_file_path = path + "/" + new_file_name + "." + file_extension
        
        scr = folder_template_location + "/" + file_name + "." + file_extension
        try:
            shutil.copyfile(scr, new_file_path)
        except:
            if not os.path.exists(new_file_path):
                return "A folder does not exist"
            else:
                return "Error"
        
        workbook = openpyxl.load_workbook(new_file_path)
        sheet = workbook[sheet_name]
        
        sheet.cell(4, 3).value = xy_input_unit
        
        for row in range(len(X)):
            sheet.cell(row+7, 3).value = X[row]
            sheet.cell(row+7, 4).value = Y[row]
            
            if len(pad_number_list) == len(X):
                sheet.cell(row+7, 2).value = pad_number_list[row]
            if len(pad_name_list) == len(X):
                sheet.cell(row+7, 5).value = pad_name_list[row]
            if len(dut_name_list) == len(X):
                sheet.cell(row+7, 1).value = dut_name_list[row]
            else:
                sheet.cell(row+7, 1).value = dut_name_format + ".0"
        
        sheet = workbook["Revision"]
        user_name = os.getlogin()
        
        columnCount = sheet.max_column
        rowCount = sheet.max_row
        
        for row in range(1, rowCount + 1):
            for column in range(1, columnCount + 1):
                if sheet.cell(row, column).value == "ORG":
                    row_index = row
                    column_index = column
                    break
        
        for row in range(row_index, rowCount + 1):
            a = sheet.cell(row, column_index).value
            b = sheet.cell(row+1, column_index).value
            if a != None and b == None:
                sheet.cell(row+1, column_index).value = os.getlogin()
            
            a = sheet.cell(row, column_index-2).value
            if a == "REVISION":
                sheet.cell(row+1, column_index - 2).value = 1
            else:
                sheet.cell(row+1, column_index - 2).value = sheet.cell(row, column_index - 2).value + 1
                           
            a = sheet.cell(row, column - 1).value
            if a == "DESCRIPTION":
                sheet.cell(row + 1, column - 1).value = "Initial"
            else:
                sheet.cell(row + 1, column - 1).value = "Update"
            
            now = datetime.now()
            sheet.cell(row+1, column_index+2).value = now.strftime("%b-%d-%Y")
            break
        
        workbook.save(new_file_path)
        return 'Done'
    else: # file_exist == True
        pass
    
    return "EXPORT_ARRAY_FULL_SITE_FOR_REFERENCE_FILE"

def EXPORT_IUA_PLUS_FILE():
    return "EXPORT_IUA_PLUS_FILE"

def EXPORT_CRD_PLUS_FILE():
    return "EXPORT_CRD_PLUS_FILE"

def EXPORT_PROBE_HEAD_XY_COORDINATES_FOR_APPROVAL_FILE(X, \
                                                        Y, \
                                                        pad_number_list, \
                                                        pad_name_list, \
                                                        dut_name_list, \
                                                        xy_input_unit,  \
                                                        status = [], \
                                                        dut_name_format="", \
                                                        card_part_number="", \
                                                        nc_list="", \
                                                        customer_name="", \
                                                        device_name=""):
                                                
    # Check X and Y list are empty or not
    if not X or not Y: # not export anything of X and Y list are empty
        # print("empty")
        return "Empty"
    
    # Declare some variables
    path = ""
    file_exist = False
    XY_exist_name = ""

    # Get file_name and file_extension from template
    file_name, file_extension = GET_FILE_PATH('paths.txt', 'PROBE_HEAD_XY_TEMPLATE_FILENAME')

    # Check file does exited or not yet base on card_part_number
    if card_part_number != "":
        path = GET_FOLDER_PATH('paths.txt', 'design_folder_path') + "/" + str(card_part_number) + '-xx'
        if os.path.exists(path):
            dir_list = os.listdir(path)
            for i in dir_list:
                string = card_part_number + "-xx_Probe Head XY Coordinates For Approval" 
                # string is like: "PCX-000000-xx_Probe Head XY Coordinates For Approval"
                if (i.find(string) != -1) and (i[i.find(".")+1:] == file_extension):  #if string is included in files list and match file extension
                    file_exist = True
                    XY_exist_name = i # get exist file name
                else: # if string is not included in files list and not match file extension
                    continue
            # end for
        else:
            return "Card Folder doesn't exist!" # end if
    else:
        return "Card part number is empty!"  # alway need card_part_number to add to file, return empty char if this variable is ""

    if file_exist == False: # if file doesn't exist, creating new file Rev01 from template Rev00
        folder_template_location    = GET_FOLDER_PATH("paths.txt", 'folder_template_path')
        revision_string_format      = GET_REVISION_STRING('paths.txt', 'REVISION_FORMAT')
        revision_pos_in_file_name   = file_name.index(revision_string_format)
        revision_string             = file_name[revision_pos_in_file_name:] # example: Rev00 from template file
        new_file_path, new_file_name = "", ""
        
        if card_part_number != "":
            new_file_name = file_name.replace("YYY-XXXXXX", card_part_number)
            new_file_name = new_file_name.replace(new_file_name[revision_pos_in_file_name:], "Rev01")
            new_file_path = path + "/" + new_file_name + "." + file_extension

        scr = folder_template_location + "/" + file_name + "." + file_extension
        try:
            shutil.copyfile(scr, new_file_path)
        except:
            if not os.path.exists(new_file_path):
                return "A folder does not exist"
            else:
                return "Error"
        print(new_file_path)
        # Load workbook
        workbook = openpyxl.load_workbook(new_file_path)
        # Active sheet 
        sheet = workbook[PROBE_HEAD_XY_SHEETS_NAME[1]]
    
        length = len(X)
        length_dut = len(dut_name_list)
        length_pad_name = len(pad_name_list)
        length_pad_number = len(pad_number_list)
        length_nc_list = len(nc_list)
        # Export data to excel file.
        for row in range(length):
            sheet.cell(row+8, 3).value = X[row]
            sheet.cell(row+8, 4).value = Y[row]
            if length_dut == length:
                sheet.cell(row+8, 1).value = dut_name_list[row]
            if length_pad_name == length:
                sheet.cell(row+8, 5).value = pad_name_list[row]
            if length_pad_number == length:
                sheet.cell(row+8, 2).value = pad_number_list[row]
            if length_nc_list == length:
                sheet.cell(row+8, 6).value = nc_list[row]
        
        sheet.cell(1,2).value = customer_name
        sheet.cell(2,2).value = device_name
        sheet.cell(3,2).value = card_part_number
        sheet.cell(4,2).value = "01"
        
        now = datetime.now()
        sheet.cell(5,2).value = now.strftime("%b-%d-%Y")

        # Active "History of Revision" sheet
        sheet = workbook[PROBE_HEAD_XY_SHEETS_NAME[0]]
        # Update data
        sheet.cell(2,1).value = sheet.cell(5,2).value
        sheet.cell(3, 2).value = os.getlogin()

        # Save workbook
        workbook.save(new_file_path)
    else: # if file does exited
        return 

    return "PROBE_HEAD_XY_COORDINATES_FOR_APPROVAL_FILE"
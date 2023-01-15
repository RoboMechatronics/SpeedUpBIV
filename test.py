import openpyxl
import numpy as np
import os
from datetime import datetime
path = "D:/Robo_data/Python/Projects/myproject2/app/templates/YYY-XXXXXX-xx_Array full sites for reference_Rev00.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook['Revision']

columnCount = sheet.max_column
rowCount    = sheet.max_row
row_index = 0
column_index = 0

for row in range(1, rowCount+1):
    for column in range(1,columnCount+1):
        if sheet.cell(row, column).value == "ORG":
            row_index = row
            column_index = column
            break

for row in range(row_index, rowCount+1):
    if sheet.cell(row, column_index).value != None and sheet.cell(row+1, column_index).value == None:
        sheet.cell(row+1, column_index).value = os.getlogin() + str(row) 
        
        if sheet.cell(row, column_index - 2).value == "REVISION":
            sheet.cell(row+1, column_index - 2).value = 1
        else:
            sheet.cell(row+1, column_index - 2).value = sheet.cell(row, column_index - 2).value + 1
        
        if sheet.cell(row, column_index - 1).value == "DESCRIPTION":
            sheet.cell(row+1, column_index - 1).value = "Initial"
        else:
            sheet.cell(row+1, column_index - 1).value = "Update"
        now = datetime.now()
        sheet.cell(row+1, column_index + 2).value = now.strftime("%b-%d-%Y")
        break
    else:
        continue

workbook.save(path)

